import os       # execução de operações como manipulação de diretórios, criação e remoção de arquivos
from openpyxl import Workbook  #biblioteca excel/  importa a classe Workbook da biblioteca openpyxl
import pdfplumber
import re
from datetime import datetime

def main():       ##sempre será executado se for chamado Main 
    try:          ## serve para tratar exessões que acontecem dentro do código 

        diretory = 'Pdf_pasta'
        Arquivos = os.listdir(diretory)   ##verifica a pasta com os pdfs
        Arquivos_quantity = len(Arquivos)   ##len retorna um numero para fazer a contagem de quantos arquivos tem

        if Arquivos_quantity == 0:   ## comparação para se for 0 pfds
            raise Exception("Não existe arquivos nessa pasta")  ##manda mensagem se não tem os arquivos 
        
        try:
            wb = Workbook()
            ws = wb.active

            ws.title = 'DADOS-Pdf'

            ws['A1'] = 'Invoice #'
            ws['B1'] = 'Date'
            ws['C1'] = 'File Name'
            ws['D1'] = 'Status'

            last_empty_line = 1   ##a contagem das linhas começa em 1
            while ws['D' + str(last_empty_line)].value  is not None: ##loop e além disso quando ele for recuperar a celula e não estiver vazio o str converte pra string
                last_empty_line += 1

        except Exception as e: 
            print(f"Não Foi Possivel Criar Arquivo Excel") 

        try: 

            for file in Arquivos: ##para um arquivo nos arquivos 
                with pdfplumber.open(diretory + "/" + file) as pdf: ##file retorna o nome do arquivo além disso ele esta buscando os pdfs 
                    Arquivos_page = pdf.pages[0]
                    pdf_text = Arquivos_page.extract_text()
                    print(pdf_text)

                inv_number_re_pattern = r'INVOICE #(\d+)'    #ele pega os carcteres do invoice
                inv_date_re_pattern = r'DATE: (\d{2}/\d{2}/\d{4})'  #ele pega os carcteres da data 

                match_number = re.search(inv_number_re_pattern, pdf_text)  #procura para mim o texto baseado na instrução
                match_date = re.search(inv_date_re_pattern, pdf_text)

                if match_number:
                    ws['A{}'.format(last_empty_line)] = match_number.group(1) # se encontrou ele retorna esse dado para mim 
                else:
                    ws['A{}'.format(last_empty_line)] = "Não consegui achar o resultado"
                
                if match_date:
                    ws['B{}'.format(last_empty_line)] = match_date.group(1)
                else:
                    ws['B{}'.format(last_empty_line)] = "Não consegui achar o resultado"
                
                ws['C{}'.format(last_empty_line)] = file
                ws['D{}'.format(last_empty_line)] = "Finalizado"

                last_empty_line += 1   #a gente garante que a proxima linha que ele pegar vai estar vazia 
        
        except Exception as e: 
            print(f"Não Foi Possivel Criar Arquivo Excel: {e}")
            ws['A{}'.format(last_empty_line)] = "Exception: {}".format(e)
            ws['B{}'.format(last_empty_line)] = match_date.group(1)
            ws['C{}'.format(last_empty_line)] = file
            ws['D{}'.format(last_empty_line)] = "Exception: {}".format(e)

            try:
                full_now = str(datetime.now()).replace(":", "-")   ##vai pegar dia o mes e o ano de agora substitui o : por -
                dot_index = full_now.index(".")
                now = full_now[:dot_index]
                print(now)
                wb.save("Invoice -{}.xlsx".format(now))

            except Exception as e: 
                print(f"Nâo foi ´possivel acessar o arquivo") 


    except Exception as e: 
        print(f"Erro Arquivo não encontrado")       

if __name__ == "__main__":
    main()
















